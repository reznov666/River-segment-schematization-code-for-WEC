# -*- coding: UTF-8 -*-
'''
python_version:2.7
Required Python standard libraries:
    os
    traceback
    shutil

Required Python third-party libraries:
    Arcpy(ArcGIS) version:10.2
    numpy         version:1.16.6
    pandas        version:0.24.2
    openpyxl      version:2.6.4

'''
import arcpy as ap
import os
import numpy as np
import openpyxl as xl
import pandas as pd
import traceback
import shutil


def find_row(from_node, did):
    for i, node in enumerate(from_node):
        if node == did:
            return i


def build_table():
    try:
        u = 0.6  # velocity of water is defaulted to 0.6
        kcod0 = 0.07 / 86400  # K value is defaulted to 0.07
        knh3 = 0.1 / 86400
        c_to_n = 15.  # COD value to NH4N value efficient

        flow_transform = 1 #CONTROL THE FLOW:the flow inputted is average flow,
        # # if it's needed to be changed into least flow, please un-mute the next row
        # flow_transform = 8

        flow_efficient_0 = 0.0056 / flow_transform
        wqt_standard = [
            ['1', 15.0, 0.15],
            ['2', 15.0, 0.5],
            ['3', 20.0, 1.0],
            ['4', 30.0, 1.5],
            ['5', 40.0, 2.0]
        ]
        wqc = []
        wqc2 = []
        for i in wqt_standard:
            wqc.append(i[0])
            wqc2.append(i[1])

        concentration_table_COD = {'1': 15.0, '2': 15.0, '3': 20.0, '4': 30.0, '5': 40.0}
        concentration_table_NH4 = {'1': 0.15, '2': 0.5, '3': 1.0, '4': 1.5, '5': 2.0}
        # target years
        #target_years = ('2020', '2025', '2030', '2035')
        target_years = ('2020',)

# -----------------------------------------------calculation begins below--------------------------------------------
        path_in = r'E:\arcgisdata\LiuLi_CODE\SWAT_for_watersheds_songhua_2\zhi/'
        version = 0
        path_out = 'E:/arcgisdata/master1/ceshi/file_output' + str(version) + '/'
        while os.path.exists(path_out):
            version += 1
            path_out = 'E:/arcgisdata/master1/ceshi/file_output' + str(version) + '/'
        os.mkdir(path_out)

        path_out_1 = path_out + 'COD/'
        path_out_2 = path_out + 'NH4N/'
        os.mkdir(path_out_1)
        os.mkdir(path_out_2)
        for year in target_years:
            os.mkdir(os.path.join(path_out_1, year))
            os.mkdir(os.path.join(path_out_2, year))

        files = os.listdir(path_in)
        files.sort(key=lambda x: x)
        for f in files:

            # extract info from supplement
            wb = xl.load_workbook(path_in + f + '/supplement.xlsx')
            ws = wb['Sheet1']
            uID_top = str(ws["a2"].value)
            C0 = str(ws['b2'].value)
            tmp = float(ws['G2'].value)

            Q0 = ws['c2'].value / flow_transform
            hydro_site = []
            c = 2
            cursor = ws['d' + str(c)].value
            while cursor:                           #TODO
                hydro_site.append(cursor.split('/'))
                c += 1
                cursor = ws['d' + str(c)].value
            c = 2
            mc = list()
            cursor = ws['e' + str(c)].value
            while cursor:
                mc.append(cursor.split('/'))
                c += 1
                cursor = ws['e' + str(c)].value
            tail_wqt = str(ws['f2'].value)

            # do mdb operation
            ap.env.workspace = path_in + f + '/' + f + '.mdb'

            # firstly, access river info
            stream_raw = np.zeros((0, 8))
            stream_fields = ('ARCID', 'GRID_CODE', 'FROM_NODE', 'TO_NODE', 'Subbasin', 'SubbasinR', 'AreaC', 'Shape_Length')
            with ap.da.SearchCursor('WDREACH', stream_fields) as cursor:
                for row in cursor:
                    stream_raw = np.vstack((stream_raw, np.array(row)))

            stream_raw_str = list()  # change stream_raw to str
            for i, row in enumerate(stream_raw[:, :-1]):
                stream_raw_str_r = list()
                for j, value in enumerate(row):
                    stream_raw_str_r.append(str(int(value)))
                stream_raw_str_r.append(stream_raw[i, -1])
                stream_raw_str.append(stream_raw_str_r)
            from_node = []
            basin = []
            to_node = []
            basinR = []
            length = []
            for i in stream_raw_str:
                from_node.append(i[2])
                to_node.append(i[3])
                basin.append(i[4])
                basinR.append(i[5])
                length.append(i[7])

            # secondly, assess to watershed table
            watershed_fields = ('GRIDCODE', 'Shape_area')
            watershed_dic = {}
            with ap.da.SearchCursor('WDWATERSHED', watershed_fields) as cursor:
                for row in cursor:
                    watershed_dic[str(row[0])] = row[1] / 10 ** 6

            # construct calculation data_table,
            first_row = ('uID','dID','subbasin','subbasinE','utype','Q0','q1','q2','Q','COD0','CODs','t','kcod','subA','EsubA')
            uid = [uID_top]
            m = find_row(from_node, uID_top)
            did = [to_node[m]]
            while did[-1] != '0':
                uid.append(from_node[find_row(from_node, did[-1])])
                did.append(to_node[find_row(from_node, did[-1])])
            df = pd.DataFrame({ 'subbasin': uid, 'dID': did, 'uID': uid })
            utype = ['T'] * did.__len__()
            df.insert(0, 'utype', utype)

            # count extra subbasins for current sb
            # subbasinE = [[]] * uid.__len__()
            basin_total = set(from_node)
            basin_in = set(uid)
            basin_out = list(basin_total - basin_in)
            basin_in_from_out = set()
            dic = {}  # basin : row
            while basin_out:
                cursor = basin_out.pop() 
                u_node = to_node[find_row(from_node, cursor)]
                row = find_row(did, u_node) 
                if row is not None:
                    basin_in_from_out.add(cursor)
                    dic[cursor] = row + 1  
                elif u_node in basin_in_from_out:
                    dic[cursor] = dic[u_node] 
                    basin_in_from_out.add(cursor)
                else:
                    basin_out.insert(0, cursor)
            basinE = []
            for i in range(uid.__len__()):
                basinE.append([])
            for i in dic:
                basinE[dic[i]].append(i) 
            str_subbasinE = [''] * uid.__len__()
            for i, ele in enumerate(basinE):
                if ele:
                    str_subbasinE[i] = str(ele).replace('\'', '').strip('[]') 
            df.insert(0, 'subbasinE', str_subbasinE)

            Q0_list = [0] * uid.__len__()
            q1 = [0] * uid.__len__()
            q2 = [0] * uid.__len__()
            Q = [0] * uid.__len__()
            COD0 = [0] * uid.__len__()
            CODs = [0] * uid.__len__()
            t = [0] * uid.__len__()
            kcod = [0] * uid.__len__()
            subA = [0] * uid.__len__()
            EsubA = [0] * uid.__len__()

            # calculate area of subA and EsubA
            for i, ele in enumerate(uid):
                subA[i] = watershed_dic[ele]
            for i, ebasin in enumerate(basinE):
                if ebasin:
                    if ebasin.__len__() == 1:
                        EsubA[i] = watershed_dic[ebasin[0]]
                    else:
                        A = 0
                        for ee in ebasin:
                            A += watershed_dic[ee]
                        EsubA[i] = A

            # calculate flow efficient
            if 'flow_efficient' not in dir():                       #TODO
                flow_efficient = []
                flow_code_list = []
                for flow_list in hydro_site:
                    flow_code = flow_list[0]
                    flow_code_list.append(flow_code)
                    flow_percentage = flow_list[1]
                    flow_value = flow_list[2]
                    sum_basin_1 = sum(subA[:uid.index(flow_code) - 1])
                    sum_basin_2 = 0
                    for i in EsubA[:uid.index(flow_code)]:
                        if i:
                            sum_basin_2 += i
                    sum_basin_3 = subA[uid.index(flow_code)] * float(flow_percentage)
                    flow_efficient_0 = (float(flow_value) - Q0) / (sum_basin_2 + sum_basin_1 + sum_basin_3)
                    if len(flow_efficient) == 0:
                        flow_efficient.extend([flow_efficient_0]*(uid.index(flow_code)+1))
                    else:
                        flow_efficient.extend([flow_efficient_0]*(uid.index(flow_code)-uid.index(flow_code_list[-2])))
                flow_efficient.extend([flow_efficient[-1]]*(uid.__len__() - flow_efficient.__len__()))

            # calculate other columns for raw table as list
            Q0_list[0] = Q0
            for i in range(uid.__len__()):
                q1[i] = EsubA[i] * flow_efficient[i]
                q2[i] = subA[i] * flow_efficient[i]
                t[i] = length[from_node.index(uid[i])] / u
                kcod[i] = kcod0 * 1.074**(tmp - 20)  #Decay coefficient is converted according to temperature.
            Q[0] = (Q0 + q1[0] + q2[0])
            for i in range(1, uid.__len__()):
                Q0_list[i] = Q[i-1]
                Q[i] = (Q0_list[i] + q1[i] + q2[i])
            COD0[0] = concentration_table_COD[C0]
            df.insert(0, 'q1', q1)
            df.insert(0, 'q2', q2)
            df.insert(0, 'Q0', Q0_list)
            df.insert(0, 'kcod', kcod)
            df.insert(0, 'COD0', COD0)
            df.insert(0, 'Q', Q)
            df.insert(0, 'CODs', CODs)
            df.insert(0, 'subA', subA)
            df.insert(0, 'EsubA', EsubA)
            df.insert(0, 't', t)

            # insert monitoring cross info as new row into dataframe
            count = 1
            mind = -1
            for i, ele in enumerate(mc):
                location = ele[0]
                ratio = float(ele[1])
                wqt = ele[2]
                parti_line = df['uID'].tolist().index(location)
                if mind == location:
                    count += 1
                else:
                    count = 1
                mind = location
                part1 = df[:parti_line + count]
                part2 = df[parti_line + count:]
                new_line = {}
                for j in df.columns:
                    new_line[j] = df[j][part1.shape[0]-1]
                # change the last line of part1
                part1['dID'][parti_line + count - 1] = 'M' + str(i)
                part1['q2'][parti_line + count - 1] *= ratio
                part1['Q'][parti_line + count - 1] = part1['Q0'][parti_line + count - 1] + part1['q1'][parti_line + count - 1] + part1['q2'][parti_line + count - 1]
                part1['CODs'][parti_line + count - 1] = wqt_standard[wqc.index(wqt)][1]  # do cod first
                part1['t'][parti_line + count - 1] *= ratio
                part1['subA'][parti_line + count - 1] *= ratio

                # change the new line
                new_line['uID'] = 'M' + str(i)
                new_line['utype'] = 'M'
                new_line['subbasinE'] = ''
                new_line['Q0'] = part1['Q'][parti_line + count - 1]
                new_line['q1'] = 0.0
                new_line['q2'] *= 1 - ratio
                new_line['Q'] = new_line['Q0'] + new_line['q1'] + new_line['q2']
                new_line['t'] *= 1 - ratio
                new_line['subA'] *= 1 - ratio
                new_line['EsubA'] = 0.0

                # merge them
                part1 = part1.append(new_line, ignore_index=True)
                part1 = part1.append(part2, ignore_index=True)
                df = part1

            # edit the last row of entire df
            df.loc[df.shape[0] - 1, 'CODs'] = wqt_standard[wqc.index(tail_wqt)][1]
# -------------------------------------------------------Output the result----------------------------------------
            # identify which row need to be changed CODs values
            series_CODs = df['CODs']
            row_CODs = []
            for i, ele in enumerate(series_CODs):
                if ele > 0.1:
                    row_CODs.append(i)
            # if f == 'alh':   ????what this???
            #     1-1
            for year_index, year in enumerate(target_years):
                wb_stream_net = xl.Workbook()
                ws = wb_stream_net.active
                # output result to table of COD
                if year != '2020':  # change CODs to correspondent year's value
                    for i, ele in enumerate(row_CODs[:-1]):
                        df.loc[ele, 'CODs'] = concentration_table_COD[str(mc[i][year_index + 2])]

                for i, ele in enumerate(first_row):
                    ws.cell(row=1, column=i + 1).value = ele
                for j in range(15):
                    sri = df[first_row[j]]
                    for i, ele in enumerate(sri):
                        ws.cell(row=i + 2, column=j + 1).value = ele
                wb_stream_net.save(os.path.join(path_out_1, year, f + '-' + str(watershed_dic.__len__()) + '-.xlsx'))

            for year_index, year in enumerate(target_years):
                # generate df_n based on df
                df_n = df
                df_n['kcod'] = knh3 * 1.045**(tmp - 20)                #TODO Decay coefficient is converted according to temperature.
                # change the C0 to value fo NH4
                df_n.loc[0, 'COD0'] = concentration_table_NH4[C0]
                df_n.loc[len(df_n['CODs']) - 1, 'CODs'] = concentration_table_NH4[tail_wqt]
                # output result to table of NH4N
                wb_stream_net = xl.Workbook()
                ws = wb_stream_net.active

                for i, ele in enumerate(row_CODs[:-1]):
                    df_n.loc[ele, 'CODs'] = concentration_table_NH4[str(mc[i][year_index + 2])]
                for i, ele in enumerate(first_row):
                    ws.cell(row=1, column=i + 1).value = ele
                for j in range(15):
                    sri = df_n[first_row[j]]
                    for i, ele in enumerate(sri):
                        ws.cell(row=i+2, column=j+1).value = ele
                wb_stream_net.save(os.path.join(path_out_2, year, f + '-' + str(watershed_dic.__len__()) + '-nh4.xlsx'))

            print('{:^16}, is finished'.format(f))
            del flow_efficient

        for path in ['COD/2020','NH4N/2020']:
            source_folder = path_out + path
            destination_folder = path_out + path.split('/')[0]

            #
            if os.path.exists(source_folder):
                #
                for item in os.listdir(source_folder):
                    source_item = os.path.join(source_folder, item)
                    destination_item = os.path.join(destination_folder, item)

                    #
                    shutil.move(source_item, destination_item)

                #
                os.rmdir(source_folder)
            else:
                print("chage path error")





    except Exception as ex:
        #print(f)
        print(ex)
        traceback.print_exc()

if __name__ == '__main__':
    build_table()

    # Define the paths of the source folder and the target folder


